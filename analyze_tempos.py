import openpyxl
import sys

try:
    # Try to load the Tempos Excel file
    wb = openpyxl.load_workbook('tempos.xls')
    ws = wb.active

    print("=" * 80)
    print("ANÁLISE DO EXCEL DE TEMPOS (TET/R2P)")
    print("=" * 80)

    # Check date in C2 (assuming similar structure)
    print(f"\n📅 Data (C2): {ws['C2'].value}")

    # Analyze rows 13-25
    print("\n📊 ESTRUTURA DAS LINHAS (13-25):")
    print("-" * 80)
    # Print first 15 columns to identify TET and R2P
    header = "Row " + " ".join([f"{chr(64+c)}".ljust(10) for c in range(1, 16)])
    print(header)
    print("-" * 80)

    for i in range(13, 26):
        row_str = f"{i:<4}"
        for c in range(1, 16):
            val = ws.cell(row=i, column=c).value
            val_str = str(val)[:9] if val is not None else "-"
            row_str += f"{val_str:<11}"
        print(row_str)

except Exception as e:
    print(f"Erro ao ler arquivo: {e}")
    print("O arquivo pode ser formato binário antigo (.xls) não suportado pelo openpyxl.")
