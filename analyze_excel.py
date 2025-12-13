import openpyxl
from datetime import datetime

# Load the Excel file
wb = openpyxl.load_workbook('Cópia de Relatório Vendas Horárias (Restaurante) AMADORA 20 10-12-2025.xlsx')
ws = wb.active

print("=" * 80)
print("ANÁLISE DO EXCEL DE VENDAS")
print("=" * 80)

# Check date in C2
print(f"\n📅 Data (C2): {ws['C2'].value}")
print(f"   Tipo: {type(ws['C2'].value)}")

# Analyze rows 13-25
print("\n📊 ESTRUTURA DAS LINHAS (13-25):")
print("-" * 80)
print(f"{'Row':<5} {'Coluna B (Timestamp)':<30} {'D (Vendas)':<12} {'Y (Horas)':<10}")
print("-" * 80)

for i in range(13, 26):
    cell_b = ws[f'B{i}'].value
    cell_d = ws[f'D{i}'].value
    cell_y = ws[f'Y{i}'].value
    
    # Format B value
    if cell_b is None:
        b_str = "VAZIO"
    elif isinstance(cell_b, datetime):
        b_str = f"{cell_b.strftime('%Y-%m-%d %H:%M:%S')} (Date)"
    elif isinstance(cell_b, (int, float)):
        b_str = f"{cell_b} (Number)"
    else:
        b_str = f"{str(cell_b)[:25]} (String)"
    
    # Format D and Y
    d_str = f"{cell_d}" if cell_d else "-"
    y_str = f"{cell_y}" if cell_y else "-"
    
    print(f"{i:<5} {b_str:<30} {d_str:<12} {y_str:<10}")

# Find totals pattern
print("\n🎯 IDENTIFICANDO TOTAIS:")
print("-" * 80)
manha_total = 0
noite_total = 0

for i in range(13, 200):
    cell_b = ws[f'B{i}'].value
    cell_d = ws[f'D{i}'].value
    
    if not cell_b and cell_d:
        prev_b = ws[f'B{i-1}'].value
        hora = prev_b.hour if isinstance(prev_b, datetime) else 0
        
        if hora < 16:
            manha_total += cell_d
        else:
            noite_total += cell_d
            
        print(f"Linha {i}: TOTAL encontrado!")
        print(f"  → Vendas: {cell_d}")
        print(f"  → Linha anterior ({i-1}): B = {prev_b}")
        if isinstance(prev_b, datetime):
            print(f"  → Hora extraída: {prev_b.hour}h → {'MANHÃ' if prev_b.hour < 16 else 'NOITE'}")
        print()

print(f"\n📊 RESUMO:")
print(f"  Manhã (< 16h): €{manha_total:.2f}")
print(f"  Noite (≥ 16h): €{noite_total:.2f}")
print("=" * 80)
