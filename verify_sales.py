import openpyxl
from datetime import datetime

# Load the Excel file
wb = openpyxl.load_workbook('Cópia de Relatório Vendas Horárias (Restaurante) AMADORA 20 10-12-2025.xlsx')
ws = wb.active

print("=" * 80)
print("ANÁLISE DETALHADA - VERIFICANDO VENDAS POR HORA")
print("=" * 80)

manha_vendas = []
noite_vendas = []

for i in range(13, 200):
    cell_b = ws[f'B{i}'].value
    cell_d = ws[f'D{i}'].value
    
    # Só processar linhas de TOTAL (B vazio + D com valor)
    if not cell_b and cell_d:
        prev_b = ws[f'B{i-1}'].value
        
        if isinstance(prev_b, datetime):
            hora = prev_b.hour
            valor = cell_d
            
            print(f"Linha {i}: Hora {hora:02d}h → €{valor:.2f}")
            
            if hora < 16:
                manha_vendas.append((hora, valor))
            else:
                noite_vendas.append((hora, valor))

print("\n" + "=" * 80)
print("RESUMO MANHÃ (< 16h):")
print("-" * 80)
total_manha = 0
for hora, valor in manha_vendas:
    print(f"  {hora:02d}h: €{valor:.2f}")
    total_manha += valor
print(f"\n  TOTAL MANHÃ: €{total_manha:.2f}")

print("\n" + "=" * 80)
print("RESUMO NOITE (≥ 16h):")
print("-" * 80)
total_noite = 0
for hora, valor in noite_vendas:
    print(f"  {hora:02d}h: €{valor:.2f}")
    total_noite += valor
print(f"\n  TOTAL NOITE: €{total_noite:.2f}")

print("\n" + "=" * 80)
print(f"TOTAL GERAL: €{total_manha + total_noite:.2f}")
print("=" * 80)

# Verificar se há vendas em milhares
print("\n⚠️  ATENÇÃO: Verificar se valores já estão em milhares (€k)")
print(f"Se dividir por 1000:")
print(f"  Manhã: €{total_manha/1000:.2f}k")
print(f"  Noite: €{total_noite/1000:.2f}k")
