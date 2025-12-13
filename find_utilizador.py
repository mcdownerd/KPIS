import openpyxl
from datetime import datetime

# Load the Perdas Excel file
wb = openpyxl.load_workbook('Relatório de Transações de Inventário 20251211 20.xlsx')
ws = wb.active

print("=" * 80)
print("PROCURANDO COLUNA 'Utilizador'")
print("=" * 80)

# Check first row for column headers
print("\nCabeçalhos (linha 1):")
for col in range(1, 30):
    cell = ws.cell(row=1, column=col)
    if cell.value:
        print(f"  Coluna {chr(64+col)}: {cell.value}")

# Find "Perdas" rows with Utilizador
print("\n🎯 LINHAS COM 'Perdas' E UTILIZADOR:")
print("-" * 80)

for i in range(13, 70):
    cell_a = ws[f'A{i}'].value
    
    if cell_a and 'Perdas' in str(cell_a):
        # Check columns for "Utilizador"
        for col in range(1, 30):
            cell = ws.cell(row=i, column=col)
            if cell.value and isinstance(cell.value, str) and len(cell.value) <= 10:
                # Might be a username
                print(f"Linha {i}: A='{cell_a}' | Col {chr(64+col)}='{cell.value}'")
                break

print("=" * 80)
