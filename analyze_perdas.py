import openpyxl
from datetime import datetime

# Load the Perdas Excel file
wb = openpyxl.load_workbook('Relatório de Transações de Inventário 20251211 20.xlsx')
ws = wb.active

print("=" * 80)
print("ANÁLISE DO EXCEL DE PERDAS/INVENTÁRIO")
print("=" * 80)

# Check date in E2
print(f"\n📅 Data (E2): {ws['E2'].value}")
print(f"   Tipo: {type(ws['E2'].value)}")

# Analyze rows 13-25
print("\n📊 ESTRUTURA DAS LINHAS (13-25):")
print("-" * 80)
print(f"{'Row':<5} {'A (Tipo)':<30} {'Q (Timestamp)':<25} {'AB (Valor)':<10}")
print("-" * 80)

for i in range(13, 26):
    cell_a = ws[f'A{i}'].value
    cell_q = ws[f'Q{i}'].value
    cell_ab = ws.cell(row=i, column=28).value  # Column AB = 28
    
    # Format A
    a_str = str(cell_a)[:28] if cell_a else "-"
    
    # Format Q
    if isinstance(cell_q, datetime):
        q_str = cell_q.strftime('%Y-%m-%d %H:%M:%S')
    else:
        q_str = str(cell_q)[:23] if cell_q else "-"
    
    # Format AB
    ab_str = f"€{cell_ab:.2f}" if cell_ab else "-"
    
    print(f"{i:<5} {a_str:<30} {q_str:<25} {ab_str:<10}")

# Find "Perdas" rows
print("\n🎯 IDENTIFICANDO LINHAS COM 'Perdas':")
print("-" * 80)
manha_perdas = 0
noite_perdas = 0

for i in range(13, 200):
    cell_a = ws[f'A{i}'].value
    cell_q = ws[f'Q{i}'].value
    cell_ab = ws.cell(row=i, column=28).value
    
    if cell_a and 'Perdas' in str(cell_a):
        if isinstance(cell_q, datetime):
            hora = cell_q.hour
        else:
            hora = 0
        
        valor = float(cell_ab) if cell_ab else 0
        
        print(f"Linha {i}: '{cell_a}' → Hora {hora}h, Valor €{valor:.2f}")
        
        if hora < 17:  # Manhã até 17h
            manha_perdas += valor
        else:
            noite_perdas += valor

print(f"\n📊 RESUMO:")
print(f"  Manhã (< 17h): €{manha_perdas:.2f}")
print(f"  Noite (≥ 17h): €{noite_perdas:.2f}")
print("=" * 80)
